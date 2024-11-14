from openpyxl import load_workbook
from openpyxl.utils import get_column_letter #para selecionar a letra da nossa coluna

# 1- Ler pasta de trabalho e planilha
wb = load_workbook("data/barchart.xlsx")
sheet = wb["Relatório"]

# 2- Referências das linhas e colunas (é preciso informar a linha máxima e mínima para o python)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3- Incluindo Fórmula

# sheet["B6"] = "=SUM(B2:B5)"
# sheet["B6"].style = "Currency"
# wb.save("test.xlsx") -> successfully done

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    # print(letter)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row })" # o "f" serve para ir acompanhando os dados no looping
    sheet[f"{letter}{max_row+1}"].style = "Currency"


wb.save("test1.xlsx")
